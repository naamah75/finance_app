from __future__ import annotations

import argparse
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from db import init_db, replace_transaction_rules, upsert_account


MONTHLY_SHEET = "OperazioniRicorrenti"
YEARLY_SHEET = "OperazioniSingole"
CARD_SETTLEMENT_DAY = 10

HEADER_ALIASES = {
    "giorno": "day",
    "descrizione": "description",
    "importo": "amount",
    "conto": "account",
    "pagamento": "payment_method",
    "finanziaria": "provider",
    "inizio": "start_date",
    "fine": "end_date",
    "rate": "installments_total",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import transaction rules from Excel workbook")
    parser.add_argument("workbook", type=Path, help="Path to the .xlsx or .xlsm workbook")
    return parser.parse_args()


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return HEADER_ALIASES.get(text, text)


def parse_amount(value: object) -> float:
    if value is None:
        raise ValueError("Missing amount")
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    text = text.replace("€", "").replace(" ", "")
    text = text.replace(".", "").replace(",", ".")
    return float(text)


def parse_date_string(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text


def parse_optional_int(value: object) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    return int(str(value).strip())


def parse_day_value(value: object, frequency: str) -> tuple[int, int | None]:
    if value is None or value == "":
        raise ValueError("Missing day")

    if isinstance(value, datetime):
        if frequency == "monthly":
            return value.day, None
        return value.day, value.month
    if isinstance(value, date):
        if frequency == "monthly":
            return value.day, None
        return value.day, value.month

    if isinstance(value, (int, float)):
        numeric_day = int(value)
        if frequency == "monthly":
            return numeric_day, None
        raise ValueError("Yearly rules require day/month values")

    text = str(value).strip()
    if frequency == "monthly":
        return int(float(text)), None

    day_text, month_text = text.split("/", maxsplit=1)
    return int(day_text), int(month_text)


def parse_day_from_start_date(value: object, frequency: str) -> tuple[int, int | None] | None:
    if value in (None, ""):
        return None
    parsed = parse_date_string(value)
    if not parsed:
        return None
    parsed_date = date.fromisoformat(parsed)
    if frequency == "monthly":
        return parsed_date.day, None
    return parsed_date.day, parsed_date.month


def normalize_payment_method(value: object) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip().lower()
    if text in {"conto", "carta"}:
        return text
    return text


def row_to_mapping(row: tuple[object, ...], headers: list[str]) -> dict[str, object]:
    return {
        headers[index]: row[index]
        for index in range(min(len(headers), len(row)))
        if headers[index]
    }


def extract_rules(workbook_path: Path) -> list[dict]:
    workbook = load_workbook(workbook_path, data_only=True)
    rules: list[dict] = []
    accounts_by_name: dict[str, int] = {}

    for account_name in ("Unicredit", "Fineco"):
        accounts_by_name[account_name] = upsert_account(account_name)

    for sheet_name, frequency in ((MONTHLY_SHEET, "monthly"), (YEARLY_SHEET, "yearly")):
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Missing worksheet: {sheet_name}")

        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [normalize_header(value) for value in rows[0]]

        for row in rows[1:]:
            mapping = row_to_mapping(row, headers)
            description = str(mapping.get("description") or "").strip()

            if not description:
                continue
            if mapping.get("amount") in (None, ""):
                continue
            if not mapping.get("account"):
                continue

            account_name = str(mapping["account"]).strip()
            if account_name not in accounts_by_name:
                accounts_by_name[account_name] = upsert_account(account_name)

            if mapping.get("day") in (None, ""):
                fallback_day = parse_day_from_start_date(mapping.get("start_date"), frequency)
                if fallback_day is None:
                    continue
                day_of_month, month_of_year = fallback_day
            else:
                day_of_month, month_of_year = parse_day_value(mapping["day"], frequency)
            payment_method = normalize_payment_method(mapping.get("payment_method"))

            rules.append(
                {
                    "account_id": accounts_by_name[account_name],
                    "description": description,
                    "amount": parse_amount(mapping["amount"]),
                    "frequency": frequency,
                    "day_of_month": day_of_month,
                    "month_of_year": month_of_year,
                    "payment_method": payment_method,
                    "provider": (str(mapping.get("provider")).strip() or None)
                    if mapping.get("provider") not in (None, "")
                    else None,
                    "start_date": parse_date_string(mapping.get("start_date")),
                    "end_date": parse_date_string(mapping.get("end_date")),
                    "installments_total": parse_optional_int(mapping.get("installments_total")),
                    "card_settlement_day": CARD_SETTLEMENT_DAY if payment_method == "carta" else None,
                    "source_sheet": sheet_name,
                }
            )

    return rules


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook

    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Only .xlsx and .xlsm workbooks are supported")
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    init_db()
    rules = extract_rules(workbook_path)
    replace_transaction_rules(rules)

    print(f"Imported {len(rules)} rules from {workbook_path.name}")


if __name__ == "__main__":
    main()
